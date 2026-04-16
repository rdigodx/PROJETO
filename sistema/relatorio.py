from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path


def criar_planilha(nome_arquivo, cor, lista_alunos):
    planilha = Workbook()
    planilha_ativa = planilha.active
    planilha_ativa.title = "ALUNOS"

    cabecalho = ["Nome completo", "Nota 1", "Nota 2", "Média", "Status"]
    planilha_ativa.append(cabecalho)

    fonte_negrito = Font(bold=True)
    alinhamento = Alignment(horizontal="center", vertical="center")
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    preenchimento = PatternFill(start_color=cor, end_color=cor, fill_type="solid")

    for col in range(1, len(cabecalho) + 1):
        cell = planilha_ativa.cell(row=1, column=col)
        cell.font = fonte_negrito
        cell.alignment = alinhamento
        cell.border = borda
        cell.fill = preenchimento

    for cadastro in lista_alunos:
        linha = [cadastro['aluno'], cadastro['n1'], cadastro['n2'], cadastro['media'], cadastro['status']]
        planilha_ativa.append(linha)

    for row in planilha_ativa.iter_rows(min_row=2, max_row=planilha_ativa.max_row, max_col=5):
        for cell in row:
            cell.alignment = alinhamento
            cell.border = borda

    colunas = ["A", "B", "C", "D", "E"]
    larguras = [25, 10, 10, 10, 12]
    for col, largura in zip(colunas, larguras):
        planilha_ativa.column_dimensions[col].width = largura

    pasta_dados = Path(__file__).resolve().parent.parent / "dados"
    pasta_dados.mkdir(exist_ok=True)

    caminho_arquivo = pasta_dados / nome_arquivo
    planilha.save(caminho_arquivo)

    return str(Path("dados") / nome_arquivo)
